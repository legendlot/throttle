#!/usr/bin/env python3
"""Reads back the workbook written by verify-xlsx.mjs using openpyxl.

openpyxl is a strict, independent OOXML reader — if it opens the file and the
cells carry the right TYPES, Excel will too. Type is the thing to assert, not
just the value: an EAN that arrives as a float has already been corrupted even
though it "looks" present.

  node apps/odo/scripts/verify-xlsx.mjs && python3 apps/odo/scripts/verify_xlsx.py
"""
import sys
import zipfile
import openpyxl

PATH = "/tmp/odo-xlsx-verify.xlsx"
fails = []


def check(cond, msg):
    if cond:
        print(f"  ok   {msg}")
    else:
        print(f"  FAIL {msg}")
        fails.append(msg)


# 1. Valid ZIP with all six expected parts.
with zipfile.ZipFile(PATH) as z:
    assert z.testzip() is None, "ZIP CRC check failed"
    names = set(z.namelist())
print(f"ZIP: {len(names)} parts, CRCs valid")
for part in ("[Content_Types].xml", "_rels/.rels", "xl/workbook.xml",
             "xl/_rels/workbook.xml.rels", "xl/styles.xml",
             "xl/worksheets/sheet1.xml"):
    check(part in names, f"part present: {part}")

# 2. openpyxl opens it and sees the right shape.
wb = openpyxl.load_workbook(PATH)
ws = wb.active
print(f"\nopenpyxl: sheet={ws.title!r} dims={ws.dimensions} max_row={ws.max_row} max_col={ws.max_column}")
check(ws.title == "Odo Export", "sheet name preserved")
check(ws.max_row == 8, f"8 rows (1 header + 7 data), got {ws.max_row}")

hdr = [c.value for c in ws[1]]
print(f"header: {hdr}")
check(hdr[0] == "channel", "first header is 'channel'")
check("ean" in hdr, "sparse-row column 'ean' still present (union of keys)")

def cell(row, col_name):
    return ws.cell(row=row, column=hdr.index(col_name) + 1)

# 3. The cases that actually matter.
ean = cell(3, "ean")
check(isinstance(ean.value, str) and ean.value == "5949998565748",
      f"13-digit EAN stayed TEXT: {ean.value!r} ({type(ean.value).__name__})")

u = cell(4, "units")
check(isinstance(u.value, (int, float)) and float(u.value) == 364,
      f"numeric string '364' became a NUMBER: {u.value!r} ({type(u.value).__name__})")

nr = cell(4, "net_revenue")
check(isinstance(nr.value, (int, float)) and abs(float(nr.value) - 48120.75) < 1e-9,
      f"numeric string '48120.75' became a NUMBER: {nr.value!r}")

pc = cell(5, "product_code")
check(isinstance(pc.value, str) and pc.value == "007",
      f"leading-zero '007' stayed TEXT: {pc.value!r} ({type(pc.value).__name__})")

esc = cell(6, "channel")
check(esc.value == 'Q&C <test> "quoted"', f"XML chars round-tripped: {esc.value!r}")

neg = cell(6, "net_revenue")
check(isinstance(neg.value, (int, float)) and float(neg.value) == -5.25,
      f"negative number: {neg.value!r}")

sparse = cell(7, "product_code")
check(sparse.value is None, f"sparse cell is empty, not shifted: {sparse.value!r}")
check(cell(7, "channel").value == "Firstcry", "sparse row's other cells align correctly")

ctrl = cell(8, "channel")
check(ctrl.value == "Credbell", f"control char stripped: {ctrl.value!r}")

# 4. Header is bold and the pane is frozen — the reasons to prefer xlsx over csv.
check(ws[1][0].font.bold is True, "header row is bold")
check(ws.freeze_panes == "A2", f"header row frozen: {ws.freeze_panes}")

print()
if fails:
    print(f"{len(fails)} CHECK(S) FAILED")
    sys.exit(1)
print("ALL CHECKS PASSED — openpyxl accepts the workbook and every type is correct")
